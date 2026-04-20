from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from typing import Any, List, Optional

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

app = FastAPI(title="Teamwork Projects Import (Excel)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Models ---


class SubTask(BaseModel):
    nombre: str
    horas: Optional[float] = 0
    responsable: Optional[str] = ""
    estado: Optional[str] = "Pendiente"
    observaciones: Optional[str] = ""
    subtareas: Optional[List["SubTask"]] = []


class Task(BaseModel):
    nombre: str
    horas: Optional[float] = 0
    responsable: Optional[str] = ""
    estado: Optional[str] = "Pendiente"
    observaciones: Optional[str] = ""
    subtareas: Optional[List[SubTask]] = []


class Project(BaseModel):
    nombre: str
    cliente: str
    fecha_inicio: str  # YYYY-MM-DD
    estimado_por: Optional[str] = ""
    tareas: List[Task] = []


# --- Teamwork Excel (same layout as official import samples) ---

HEADERS = (
    "TASKLIST",
    "TASK",
    "DESCRIPTION",
    "ASSIGN TO",
    "START DATE",
    "DUE DATE",
    "PRIORITY",
    "ESTIMATED TIME",
    "TAGS",
    "STATUS",
)


def _excel_date(d: date) -> datetime:
    return datetime(d.year, d.month, d.day)


def _format_estimated_time(hours: float) -> str:
    if hours is None or hours <= 0:
        return ""
    total_min = int(round(float(hours) * 60))
    if total_min <= 0:
        return ""
    if total_min < 60:
        return f"{total_min}m"
    h, m = divmod(total_min, 60)
    if m == 0:
        return f"{h}h"
    return f"{h}h {m}m"


def _map_status(estado: Optional[str]) -> str:
    s = (estado or "").strip()
    if s.upper() == "DONE" or s.lower() == "completado":
        return "Completed"
    if s == "Bloqueado":
        return "Deferred"
    return "Active"


def _priority_for(estado: Optional[str]) -> str:
    if (estado or "") == "Bloqueado":
        return "High"
    return "Medium"


def _clean_activity_title(name: str) -> str:
    """Strip leading '-' from stored titles; hierarchy dashes are generated on export."""
    s = (name or "").strip()
    while s.startswith("-"):
        s = s[1:].lstrip()
    return s.strip()


def _task_cell(depth: int, nombre: str) -> str:
    """
    depth 0 = nombre del proyecto (sin guiones).
    depth 1 = -Sección (Preventa, DESARROLLO, …).
    depth 2+ = --, ---, … para niveles inferiores.
    """
    body = _clean_activity_title(nombre) or (nombre or "").strip()
    if depth <= 0:
        return body
    return ("-" * depth) + body


def _append_scheduled_leaf(
    rows: list[list[Any]],
    cliente: str,
    tag_proyecto: Optional[str],
    base: date,
    day_cursor: int,
    task_cell: str,
    desc: Optional[str],
    assign: Optional[str],
    hours: float,
    estado: str,
) -> int:
    hours = float(hours or 0)
    estado = estado or "Pendiente"
    if hours > 0:
        days_block = max(1, int((hours + 7.999) // 8))
        d_start = base + timedelta(days=day_cursor)
        d_end = d_start + timedelta(days=days_block - 1)
        day_cursor += days_block
    else:
        d_start = base + timedelta(days=day_cursor)
        d_end = d_start

    rows.append(
        [
            cliente,
            task_cell,
            desc,
            assign,
            _excel_date(d_start),
            _excel_date(d_end),
            _priority_for(estado),
            _format_estimated_time(hours),
            tag_proyecto,
            _map_status(estado),
        ]
    )
    return day_cursor


def _emit_subtree_rows(
    rows: list[list[Any]],
    cliente: str,
    tag_proyecto: Optional[str],
    base: date,
    day_cursor: int,
    sub: SubTask,
    depth: int,
) -> int:
    """Subtareas bajo una sección: depth 2 = --, 3 = ---, …"""
    task_cell = _task_cell(depth, sub.nombre)
    desc = (sub.observaciones or "").strip() or None
    assign = (sub.responsable or "").strip() or None
    children = sub.subtareas or []

    if children:
        rows.append(
            [
                cliente,
                task_cell,
                desc,
                assign,
                None,
                None,
                None,
                "",
                tag_proyecto,
                None,
            ]
        )
        dc = day_cursor
        for ch in children:
            dc = _emit_subtree_rows(rows, cliente, tag_proyecto, base, dc, ch, depth + 1)
        return dc

    return _append_scheduled_leaf(
        rows,
        cliente,
        tag_proyecto,
        base,
        day_cursor,
        task_cell,
        desc,
        assign,
        float(sub.horas or 0),
        sub.estado or "Pendiente",
    )


def _emit_top_level_section(
    rows: list[list[Any]],
    cliente: str,
    tag_proyecto: Optional[str],
    base: date,
    day_cursor: int,
    task: Task,
) -> int:
    """Sección de primer nivel bajo el proyecto: -Preventa, -DESARROLLO, …"""
    cell = _task_cell(1, task.nombre)
    subs = task.subtareas or []

    if not subs:
        return _append_scheduled_leaf(
            rows,
            cliente,
            tag_proyecto,
            base,
            day_cursor,
            cell,
            (task.observaciones or "").strip() or None,
            (task.responsable or "").strip() or None,
            float(task.horas or 0),
            task.estado or "Pendiente",
        )

    rows.append(
        [
            cliente,
            cell,
            (task.observaciones or "").strip() or None,
            None,
            None,
            None,
            None,
            "",
            tag_proyecto,
            None,
        ]
    )
    dc = day_cursor
    for sub in subs:
        dc = _emit_subtree_rows(rows, cliente, tag_proyecto, base, dc, sub, depth=2)
    return dc


def teamwork_import_rows(project: Project) -> list[list[Any]]:
    """
    TASKLIST = cliente en todas las filas.
    Primera fila de datos: TASK = nombre del proyecto (sin '-').
    Luego -Sección (-- ítems bajo cada sección, --- bajo subniveles, etc.).
    TAGS = nombre del proyecto (referencia).
    """
    rows: list[list[Any]] = [list(HEADERS)]
    base = datetime.strptime(project.fecha_inicio, "%Y-%m-%d").date()
    day_cursor = 0
    cliente = (project.cliente or "").strip() or "Cliente"
    proj_title = (project.nombre or "").strip() or "Proyecto"
    tag_proyecto = proj_title

    root_desc_parts: list[str] = []
    if (project.estimado_por or "").strip():
        root_desc_parts.append(f"Estimado por: {project.estimado_por.strip()}")
    root_desc = "\n".join(root_desc_parts) if root_desc_parts else None

    rows.append(
        [
            cliente,
            _task_cell(0, proj_title),
            root_desc,
            None,
            None,
            None,
            None,
            "",
            tag_proyecto,
            None,
        ]
    )

    for task in project.tareas:
        day_cursor = _emit_top_level_section(
            rows, cliente, tag_proyecto, base, day_cursor, task
        )

    return rows


def generate_teamwork_xlsx(project: Project) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in teamwork_import_rows(project):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _serialize_cell(v: Any) -> Any:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


# --- Import: Excel «Estimación Desarrollo RPA» (Rocketbot / portal) ---


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v).strip()


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").lower())


def _parse_hours_cell(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    t = str(v).strip().replace(",", ".")
    if not t or t.lower() == "horas":
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _map_estado_desarrollo_excel(s: str) -> str:
    u = (s or "").strip().upper()
    if u in ("PENDING", "PENDIENTE"):
        return "Pendiente"
    if u in ("DONE", "COMPLETED", "COMPLETADO"):
        return "Completado"
    if u in ("BLOCKED", "BLOQUEADO", "DEFERRED"):
        return "Bloqueado"
    if u == "ACTIVE":
        return "Pendiente"
    return "Pendiente"


def _b_matches_module_title(module_a: str, b: str) -> bool:
    """Si B repite el título del ítem (sin prefijo numérico), no creamos carpeta intermedia."""
    a = _norm_key(module_a)
    b = _norm_key(b)
    if not a or not b:
        return False
    a_nop = re.sub(r"^\d+\s*-\s*", "", a).strip()
    if b == a_nop or b in a or a_nop == b:
        return True
    return False


def _new_subtask_leaf(nombre: str, horas: float, estado_raw: str, obs: str) -> SubTask:
    return SubTask(
        nombre=(nombre or "").strip() or "Sin nombre",
        horas=float(horas or 0),
        responsable="",
        estado=_map_estado_desarrollo_excel(estado_raw),
        observaciones=_cell_str(obs),
        subtareas=[],
    )


def _new_subtask_group(nombre: str) -> SubTask:
    return SubTask(
        nombre=(nombre or "").strip() or "Grupo",
        horas=0,
        responsable="",
        estado="Pendiente",
        observaciones="",
        subtareas=[],
    )


def _find_desarrollo_estimacion_sheet(wb: Any) -> Any:
    """Primera hoja que tenga el bloque de detalle (formato estándar del libro de estimación)."""
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, min(ws.max_row + 1, 80)):
            a = _cell_str(ws.cell(r, 1).value)
            nk = _norm_key(a)
            if "detalle" in nk and "estim" in nk:
                return ws
    raise ValueError(
        "No se encontró una hoja con «Detalle de la estimación». "
        "El archivo debe seguir el formato del libro de estimación de desarrollo RPA."
    )


def _extract_desarrollo_sheet_metadata(ws: Any) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for r in range(1, min(ws.max_row + 1, 25)):
        a = _cell_str(ws.cell(r, 1).value)
        b = _cell_str(ws.cell(r, 2).value)
        if not a:
            continue
        al = a.lower().rstrip(":")
        if al.startswith("cliente"):
            out["cliente"] = b
        elif al.startswith("proceso"):
            out["proceso"] = b
        elif "estimado por" in al:
            out["estimado_por"] = b
        elif "estimación total" in al and "horas" in al:
            try:
                out["estimacion_total_horas"] = float(str(b).replace(",", "."))
            except ValueError:
                out["estimacion_total_horas"] = b
    return out


def parse_desarrollo_estimacion_sheet(ws: Any) -> tuple[list[SubTask], dict[str, Any]]:
    """
    Lee la tabla bajo «Detalle de la estimación»: columnas A=Item-aplicativo, B=Actividad,
    C=Horas, D=Estado, E=Observaciones (como en «Copia de Estimación Desarrollo_*.xlsx»).
    Devuelve árbol de SubTask (cada ítem-aplicativo = nodo raíz del import) y metadatos de cabecera.
    """
    meta = _extract_desarrollo_sheet_metadata(ws)
    detail_row: Optional[int] = None
    for r in range(1, min(ws.max_row + 1, 80)):
        a = _cell_str(ws.cell(r, 1).value)
        nk = _norm_key(a)
        if "detalle" in nk and "estim" in nk:
            detail_row = r
            break
    if detail_row is None:
        raise ValueError("No se encontró la fila «Detalle de la estimación».")

    i = detail_row + 1
    while i <= ws.max_row:
        a = _cell_str(ws.cell(i, 1).value)
        b = _cell_str(ws.cell(i, 2).value)
        if a.lower() == "item-aplicativo" and "actividad" in b.lower():
            i += 1
            break
        i += 1

    roots: list[SubTask] = []
    current_module: Optional[SubTask] = None
    current_folder: Optional[SubTask] = None

    def append_container_leaf(leaf: SubTask) -> None:
        if current_folder is not None:
            current_folder.subtareas.append(leaf)
        elif current_module is not None:
            current_module.subtareas.append(leaf)
        else:
            roots.append(leaf)

    while i <= ws.max_row:
        a = _cell_str(ws.cell(i, 1).value)
        b = _cell_str(ws.cell(i, 2).value)
        c_raw = ws.cell(i, 3).value
        d = _cell_str(ws.cell(i, 4).value)
        e = _cell_str(ws.cell(i, 5).value)
        hours = _parse_hours_cell(c_raw)
        c_str = _cell_str(c_raw).lower()

        if not a and not b:
            i += 1
            continue

        if a.lower() == "item-aplicativo":
            i += 1
            continue

        if a:
            current_module = _new_subtask_group(a)
            roots.append(current_module)
            current_folder = None

            if c_str == "horas" and hours is None:
                i += 1
                continue

            if hours is not None and b:
                nombre = f"{a} — {b}" if b != a else b or a
                current_module.subtareas.append(_new_subtask_leaf(nombre, hours, d, e))
                i += 1
                continue

            if b and hours is None:
                if not _b_matches_module_title(a, b):
                    current_folder = _new_subtask_group(b)
                    current_module.subtareas.append(current_folder)
                else:
                    current_folder = None
            i += 1
            continue

        if c_str == "horas" and hours is None:
            i += 1
            continue

        if not b:
            i += 1
            continue

        if hours is not None:
            append_container_leaf(_new_subtask_leaf(b, hours, d, e))
            i += 1
            continue

        if current_module is None:
            current_module = _new_subtask_group("(sin ítem-aplicativo)")
            roots.append(current_module)
        current_folder = _new_subtask_group(b)
        current_module.subtareas.append(current_folder)
        i += 1

    return roots, meta


# --- Routes ---


@app.post("/import-desarrollo-xlsx")
async def import_desarrollo_xlsx(file: UploadFile = File(...)):
    """
    Importa actividades de la etapa Desarrollo desde el Excel estándar
    («Estimación de desarrollo RPA» con tabla Detalle / Item-aplicativo).
    """
    fn = (file.filename or "").lower()
    if not fn.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Se requiere un archivo .xlsx")
    try:
        raw = await file.read()
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = _find_desarrollo_estimacion_sheet(wb)
        subtareas, meta = parse_desarrollo_estimacion_sheet(ws)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e

    return {
        "subtareas": [s.model_dump() for s in subtareas],
        "metadata": meta,
    }


@app.post("/preview-excel-data")
async def preview_excel_data(project: Project):
    raw = teamwork_import_rows(project)
    return {
        "rows": [[_serialize_cell(c) for c in r] for r in raw[:40]],
    }


@app.post("/generate-xlsx")
async def generate_xlsx(project: Project):
    try:
        data = generate_teamwork_xlsx(project)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e)) from e

    safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in project.nombre.strip())
    filename = f"{safe or 'proyecto'}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/", response_class=HTMLResponse)
async def index():
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()
